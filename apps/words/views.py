"""
Word CRUD views with role-based permissions.
"""
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import Q
from django.conf import settings
from django.http import FileResponse
from rest_framework.exceptions import NotFound
from django.utils.dateparse import parse_datetime
from django.utils import timezone

from .models import Word, CheckerAction
from .serializers import (
    WordSerializer,
    WordCreateSerializer,
    WordUpdateSerializer,
    WordUploadSerializer,
    MyWordsSerializer,
    _parse_list_string,
)
from .permissions import IsWordUploader, IsWordChecker, IsWordManager
from core.utils import success_response


class WordListCreateView(APIView):
    """List words (with search/filter) and upload word."""
    permission_classes = [IsAuthenticated, IsWordUploader]

    def get(self, request):
        """List words with search and filter by verified/unverified."""
        qs = Word.objects.all().select_related('uploaded_by', 'verified_by')
        search = request.query_params.get('search')
        is_verified = request.query_params.get('is_verified')
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))

        if search:
            qs = qs.filter(
                Q(word__icontains=search) |
                Q(meaning__icontains=search)
            )
        if is_verified is not None:
            if is_verified.lower() in ('true', '1', 'yes'):
                qs = qs.filter(is_verified=True)
            elif is_verified.lower() in ('false', '0', 'no'):
                qs = qs.filter(is_verified=False)

        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        words = qs[start:end]
        serializer = WordSerializer(words, many=True)

        return Response(success_response({
            'items': serializer.data,
            'total': total,
            'page': page,
            'page_size': page_size,
        }, "Words retrieved successfully"))

    def post(self, request):
        """Upload a new word (word_uploader, word_checker, word_manager, admin)."""
        serializer = WordCreateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            word = Word.objects.get(pk=serializer.instance.pk)
            return Response(
                success_response(WordSerializer(word).data, "Word uploaded successfully"),
                status=status.HTTP_201_CREATED
            )
        return Response(
            {"success": False, "message": "Validation failed", "data": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )


class WordsByClassChapterView(APIView):
    """GET /api/words/by-class-chapter/?class=5&chapter=3 - words grouped by length."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        class_val = request.query_params.get('class')
        chapter_val = request.query_params.get('chapter')
        if not class_val or not chapter_val:
            return Response({'1': [], '2': [], '3': [], '4': [], '5': []})
        class_val = str(class_val)
        chapter_val = str(chapter_val)
        all_words = Word.objects.all()
        matched = []
        for w in all_words:
            classes = _parse_list_string(w.class_name)
            chapters = _parse_list_string(w.chapter)
            if class_val in classes and chapter_val in chapters:
                matched.append(w)
        grouped = {'1': [], '2': [], '3': [], '4': [], '5': []}
        for w in matched:
            length_key = str(w.length) if 1 <= w.length <= 5 else '5'
            if length_key in grouped:
                grouped[length_key].append({'id': w.id, 'word': w.word})
        return Response(grouped)


class MyWordsView(APIView):
    """GET /api/words/my-words - words uploaded by current user, supports ?updated_after=<timestamp>."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Word.objects.filter(uploaded_by=request.user).order_by('-updated_at')
        updated_after = request.query_params.get('updated_after')
        if updated_after:
            dt = parse_datetime(updated_after)
            if dt:
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
                qs = qs.filter(updated_at__gt=dt)
        serializer = MyWordsSerializer(qs, many=True)
        return Response({'words': serializer.data})


class WordDetailView(APIView):
    """Retrieve, update, delete word."""
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.request.method in ('PUT', 'PATCH', 'DELETE'):
            return [IsAuthenticated(), IsWordManager()]
        return [IsAuthenticated(), IsWordUploader()]

    def get_object(self, pk):
        try:
            return Word.objects.select_related('uploaded_by', 'verified_by').get(pk=pk)
        except Word.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound("Word not found")

    def get(self, request, pk):
        """Get word by ID."""
        word = self.get_object(pk)
        serializer = WordSerializer(word)
        return Response(success_response(serializer.data, "Word retrieved successfully"))

    def put(self, request, pk):
        """Update word (word_manager or admin)."""
        word = self.get_object(pk)
        serializer = WordUpdateSerializer(word, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(success_response(WordSerializer(word).data, "Word updated successfully"))
        return Response(
            {"success": False, "message": "Validation failed", "data": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk):
        """Delete word (word_manager or admin)."""
        word = self.get_object(pk)
        word.delete()
        return Response(success_response(None, "Word deleted successfully"), status=status.HTTP_200_OK)


class WordDetailsByWordView(APIView):
    """
    GET /api/words/details?word=<word>
    Return detailed information for a single word searched by its text.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        word_text = request.query_params.get("word")
        if not word_text:
            return Response(
                {"detail": "Query parameter 'word' is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        word = (
            Word.objects.select_related("uploaded_by", "approved_by")
            .filter(word=word_text)
            .first()
        )
        if not word:
            raise NotFound("Word not found")

        class_list = _parse_list_string(word.class_name)
        chapter_list = _parse_list_string(word.chapter)

        data = {
            "word": word.word,
            "meaning": word.meaning,
            "hint1": word.hint1,
            "hint2": word.hint2,
            "class_name": class_list,
            "chapter": chapter_list,
            "added_by": getattr(word.uploaded_by, "email", None),
            "is_approved": bool(word.is_approved),
            "is_rejected": bool(word.is_rejected),
            "approved_by": getattr(word.approved_by, "email", None),
            "created_time": word.created_at.date().isoformat() if word.created_at else None,
            "updated_time": word.updated_at.date().isoformat() if word.updated_at else None,
        }
        return Response(data)


class WordVerifyView(APIView):
    """Verify word (word_checker, word_manager, admin)."""
    permission_classes = [IsAuthenticated, IsWordChecker]

    def post(self, request, pk):
        """Mark word as verified."""
        try:
            word = Word.objects.get(pk=pk)
        except Word.DoesNotExist:
            raise NotFound("Word not found")

        from django.utils import timezone
        word.is_verified = True
        word.verified_by = request.user
        word.verified_at = timezone.now()
        word.save()
        serializer = WordSerializer(word)
        return Response(success_response(serializer.data, "Word verified successfully"))


class ExcelTemplateDownloadView(APIView):
    """Download the Excel template file for word upload."""
    permission_classes = [IsAuthenticated, IsWordUploader]

    def get(self, request):
        """Serve sanswordle.xlsx (or sanswodle.xlsx) as downloadable attachment."""
        static_dir = settings.BASE_DIR / 'static'
        for name in ('sanswordle.xlsx', 'sanswodle.xlsx'):
            file_path = static_dir / name
            if file_path.exists():
                return FileResponse(
                    open(file_path, 'rb'),
                    as_attachment=True,
                    filename=name,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        raise NotFound("Excel template not found")


class ExcelUploadView(APIView):
    """Process uploaded Excel file and create words (skip if already exists)."""
    permission_classes = [IsAuthenticated, IsWordUploader]

    def post(self, request):
        """Process Excel file: create new words, skip existing, return per-row status."""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response(
                {"success": False, "message": "No file provided", "data": None},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not uploaded_file.name.lower().endswith('.xlsx'):
            return Response(
                {"success": False, "message": "Only .xlsx files are allowed", "data": None},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from openpyxl import load_workbook
        except ImportError:
            return Response(
                {"success": False, "message": "Excel support not installed", "data": None},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        try:
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"success": False, "message": "Invalid Excel file", "data": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        header_keywords = ('word', 'meaning', 'hint1', 'hint2')

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            word_val = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            if row_idx == 1 and word_val.lower() in header_keywords:
                continue  # Skip header row

            if not word_val:
                continue

            meaning_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            hint1_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            hint2_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''

            row_data = {"word": word_val, "meaning": meaning_val, "hint1": hint1_val, "hint2": hint2_val}

            if Word.objects.filter(word=word_val).exists():
                results.append({**row_data, "status": "already there"})
                continue

            serializer = WordCreateSerializer(
                data={"word": word_val, "meaning": meaning_val, "hint1": hint1_val, "hint2": hint2_val},
                context={'request': request}
            )
            if serializer.is_valid():
                serializer.save()
                results.append({**row_data, "status": "Successfully updated"})
            else:
                results.append({
                    **row_data,
                    "status": "error occurred",
                    "error": serializer.errors
                })

        wb.close()
        return Response(success_response({"results": results}, "Excel processing completed"))


class WordTemplateView(APIView):
    """Generate and return Excel template for bulk word upload. GET /api/words/template/"""
    permission_classes = [IsAuthenticated, IsWordUploader]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from io import BytesIO
        except ImportError:
            return Response(
                {"success": False, "message": "Excel support not installed"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "Words"
        headers = ['Word', 'Meaning', 'Hint1', 'Hint2', 'Class', 'Chapter']
        ws.append(headers)
        ws.append(['राम', 'Rama', 'Ayodhya king', 'Son of Dasharatha', '5', '3'])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(
            buffer,
            as_attachment=True,
            filename='word_upload_template.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


class BulkWordUploadView(APIView):
    """
    Bulk upload words from Excel file.
    POST /api/words/bulk-upload/
    Uses same logic as Word Upload: trim, duplicate check, class/chapter list updates.
    Returns status per row: uploaded, updated, error.
    """
    permission_classes = [IsAuthenticated, IsWordUploader]

    def post(self, request):
        from openpyxl import load_workbook
        from rest_framework.exceptions import ValidationError

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response(
                {"success": False, "message": "No file provided", "results": []},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not uploaded_file.name.lower().endswith('.xlsx'):
            return Response(
                {"success": False, "message": "Only .xlsx files are allowed", "results": []},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"success": False, "message": f"Invalid Excel file: {str(e)}", "results": []},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        header_keywords = ('word', 'meaning', 'hint1', 'hint2', 'class', 'chapter')

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            word_val = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            if row_idx == 1 and word_val.lower() in header_keywords:
                continue  # Skip header row

            if not word_val:
                continue

            meaning_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            hint1_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            hint2_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            class_val = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            chapter_val = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''

            row_data = {
                'word': word_val,
                'meaning': meaning_val,
                'hint1': hint1_val,
                'hint2': hint2_val,
                'class_num': class_val,
                'chapter': chapter_val,
            }

            serializer = WordUploadSerializer(data=row_data, context={'request': request})
            if not serializer.is_valid():
                err_msg = serializer.errors.get('non_field_errors', serializer.errors)
                if isinstance(err_msg, (list, dict)):
                    err_msg = err_msg[0] if isinstance(err_msg, list) and err_msg else str(err_msg)
                results.append({
                    "word": word_val,
                    "status": "error",
                    "message": str(err_msg)
                })
                continue

            try:
                serializer.save()
                if getattr(serializer, '_was_update', False):
                    results.append({
                        "word": word_val,
                        "status": "updated",
                        "message": "Class and chapter updated"
                    })
                else:
                    results.append({
                        "word": word_val,
                        "status": "uploaded",
                        "message": "Word uploaded successfully"
                    })
            except ValidationError as e:
                msg = e.detail[0] if isinstance(e.detail, list) and e.detail else str(e.detail)
                results.append({
                    "word": word_val,
                    "status": "error",
                    "message": str(msg)
                })

        wb.close()
        return Response({"results": results})


class BulkUploadView(APIView):
    """
    Bulk upload words from Excel file.
    POST /api/words/bulk-upload
    Uses same logic as Word Upload: trim, duplicate check, class/chapter updates.
    Returns per-row status: uploaded, updated, error.
    """
    permission_classes = [IsAuthenticated, IsWordUploader]

    def post(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response(
                {"success": False, "message": "No file provided", "results": []},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not uploaded_file.name.lower().endswith('.xlsx'):
            return Response(
                {"success": False, "message": "Only .xlsx files are allowed", "results": []},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            from openpyxl import load_workbook
        except ImportError:
            return Response(
                {"success": False, "message": "Excel support not installed", "results": []},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        try:
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"success": False, "message": f"Invalid Excel file: {str(e)}", "results": []},
                status=status.HTTP_400_BAD_REQUEST
            )
        required_headers = ('word', 'meaning', 'hint1', 'hint2', 'class', 'chapter')
        results = []
        header_keywords = ('word', 'meaning', 'hint1', 'hint2', 'class', 'chapter')
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            word_val = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            if row_idx == 1 and word_val.lower() in header_keywords:
                continue
            if not word_val:
                continue
            meaning_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            hint1_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            hint2_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            class_val = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            chapter_val = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            row_data = {
                'word': word_val,
                'meaning': meaning_val,
                'hint1': hint1_val,
                'hint2': hint2_val,
                'class_num': class_val,
                'chapter': chapter_val,
            }
            serializer = WordUploadSerializer(data=row_data, context={'request': request})
            if serializer.is_valid():
                try:
                    serializer.save()
                    if getattr(serializer, '_was_update', False):
                        results.append({
                            "word": word_val,
                            "status": "updated",
                            "message": "Class and chapter updated"
                        })
                    else:
                        results.append({
                            "word": word_val,
                            "status": "uploaded",
                            "message": "Word uploaded successfully"
                        })
                except Exception as e:
                    from rest_framework.exceptions import ValidationError
                    if isinstance(e, ValidationError):
                        msg = e.detail[0] if isinstance(e.detail, list) and e.detail else str(e.detail)
                    else:
                        msg = str(e)
                    results.append({"word": word_val, "status": "error", "message": msg})
            else:
                err = serializer.errors
                msg = err.get('non_field_errors', err)
                if isinstance(msg, (list, dict)):
                    msg = list(msg.values())[0][0] if isinstance(msg, dict) else (msg[0] if msg else "Validation failed")
                results.append({"word": word_val, "status": "error", "message": str(msg)})
        wb.close()
        return Response({"success": True, "results": results})


class WordTemplateView(APIView):
    """Generate and return Excel template for bulk word upload. GET /api/words/template"""
    permission_classes = [IsAuthenticated, IsWordUploader]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from io import BytesIO
        except ImportError:
            return Response(
                {"status": "error", "message": "Excel support not installed"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "Words"
        headers = ['Word', 'Meaning', 'Hint1', 'Hint2', 'Class', 'Chapter']
        ws.append(headers)
        ws.append(['राम', 'Rama', 'Ayodhya king', 'Son of Dasharatha', '5', '3'])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(
            buffer,
            as_attachment=True,
            filename='word_upload_template.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


class WordBulkUploadView(APIView):
    """Bulk upload words from Excel. POST /api/words/bulk-upload"""
    permission_classes = [IsAuthenticated, IsWordUploader]

    def post(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response(
                {"status": "error", "message": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not uploaded_file.name.lower().endswith('.xlsx'):
            return Response(
                {"status": "error", "message": "Only .xlsx files are allowed"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            from openpyxl import load_workbook
        except ImportError:
            return Response(
                {"status": "error", "message": "Excel support not installed"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        try:
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"status": "error", "message": f"Invalid Excel file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        required_headers = ('word', 'meaning', 'hint1', 'hint2')
        results = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            word_val = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            if row_idx == 1 and word_val.lower() in required_headers:
                continue
            if not word_val:
                continue
            meaning_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            hint1_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            hint2_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            class_val = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            chapter_val = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            row_data = {
                'word': word_val,
                'meaning': meaning_val,
                'hint1': hint1_val,
                'hint2': hint2_val,
                'class_num': class_val,
                'chapter': chapter_val,
            }
            serializer = WordUploadSerializer(data=row_data, context={'request': request})
            if serializer.is_valid():
                try:
                    serializer.save()
                    if getattr(serializer, '_was_update', False):
                        results.append({
                            'word': word_val,
                            'status': 'updated',
                            'message': 'Class and chapter updated'
                        })
                    else:
                        results.append({
                            'word': word_val,
                            'status': 'uploaded',
                            'message': 'Word uploaded successfully'
                        })
                except Exception as e:
                    from rest_framework.exceptions import ValidationError
                    if isinstance(e, ValidationError):
                        d = getattr(e, 'detail', None)
                        msg = d[0] if isinstance(d, list) and d else str(d) if d else str(e)
                    else:
                        msg = str(e)
                    results.append({'word': word_val, 'status': 'error', 'message': str(msg)})
            else:
                err = serializer.errors
                msg = err.get('non_field_errors', list(err.values())[0] if err else 'Validation failed')
                if isinstance(msg, (list, dict)):
                    msg = msg[0] if isinstance(msg, list) and msg else str(msg)
                results.append({'word': word_val, 'status': 'error', 'message': str(msg)})
        wb.close()
        return Response({'results': results})


class WordUploadView(APIView):
    """
    Upload a word with class and chapter.
    POST /api/words/upload
    Handles: trim, duplicate check, class/chapter list updates.
    """
    permission_classes = [IsAuthenticated, IsWordUploader]

    def post(self, request):
        data = dict(request.data)
        if 'class' in data:
            data['class_num'] = data.pop('class')
        serializer = WordUploadSerializer(data=data, context={'request': request})

        if not serializer.is_valid():
            errors = serializer.errors
            msg = errors.get('non_field_errors', errors)
            if isinstance(msg, list):
                msg = msg[0] if msg else "Validation failed"
            return Response(
                {"status": "error", "message": str(msg)},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            serializer.save()
        except Exception as e:
            from rest_framework.exceptions import ValidationError
            if isinstance(e, ValidationError):
                msg = e.detail[0] if isinstance(e.detail, list) and e.detail else str(e.detail)
                return Response(
                    {"status": "error", "message": str(msg)},
                    status=status.HTTP_400_BAD_REQUEST
                )
            raise

        if getattr(serializer, '_was_update', False):
            return Response(
                {"status": "success", "message": "Word already existed but class and chapter updated"},
                status=status.HTTP_200_OK
            )
        return Response(
            {"status": "success", "message": "Word uploaded successfully"},
            status=status.HTTP_201_CREATED
        )


# --- Check Words API (for word_checker / admin) ---

class CheckWordsRandomView(APIView):
    """
    GET /api/checkwords/random/ - assign and return one random word to check.
    A word is only assigned to one checker at a time.
    """
    permission_classes = [IsAuthenticated, IsWordChecker]

    def get(self, request):
        with transaction.atomic():
            qs = (
                Word.objects.select_for_update(skip_locked=True)
                .filter(
                    is_approved=False,
                    is_rejected=False,
                    assigned_to__isnull=True,
                )
            )
            # Avoid own uploaded words for non-admins
            if getattr(request.user, "role", None) != "admin":
                qs = qs.exclude(uploaded_by=request.user)
            word = qs.order_by("?").first()
            if not word:
                return Response(
                    {
                        "word": None,
                        "message": "No words available for checking right now.",
                    }
                )

            word.assigned_to = request.user
            word.assigned_time = timezone.now()
            word.save(update_fields=["assigned_to", "assigned_time"])

        return Response(
            {
                "id": word.id,
                "word": word.word,
                "meaning": word.meaning or "",
                "hint1": word.hint1 or "",
                "hint2": word.hint2 or "",
            }
        )


class CheckWordsApproveView(APIView):
    """POST /api/checkwords/approve/ - approve word."""
    permission_classes = [IsAuthenticated, IsWordChecker]

    def post(self, request):
        word_id = request.data.get('word_id')
        if not word_id:
            return Response({'status': 'error', 'message': 'word_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            word = Word.objects.get(pk=word_id)
        except Word.DoesNotExist:
            return Response({'status': 'error', 'message': 'Word not found'}, status=status.HTTP_404_NOT_FOUND)
        if word.assigned_to_id is not None and word.assigned_to_id != request.user.id:
            return Response(
                {'status': 'error', 'message': 'You are not assigned to this word'},
                status=status.HTTP_403_FORBIDDEN,
            )
        word.is_approved = True
        word.approved_by = request.user
        word.assigned_to = None
        word.save(update_fields=['is_approved', 'approved_by', 'assigned_to'])
        CheckerAction.objects.update_or_create(
            word=word, user=request.user,
            defaults={'action': 'approved'}
        )
        return Response({'status': 'success', 'message': 'Word approved'})


class CheckWordsUpdateApproveView(APIView):
    """POST /api/checkwords/update-approve/ - update word and approve."""
    permission_classes = [IsAuthenticated, IsWordChecker]

    def post(self, request):
        word_id = request.data.get('word_id')
        if not word_id:
            return Response({'status': 'error', 'message': 'word_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            word = Word.objects.get(pk=word_id)
        except Word.DoesNotExist:
            return Response({'status': 'error', 'message': 'Word not found'}, status=status.HTTP_404_NOT_FOUND)
        if word.assigned_to_id is not None and word.assigned_to_id != request.user.id:
            return Response(
                {'status': 'error', 'message': 'You are not assigned to this word'},
                status=status.HTTP_403_FORBIDDEN,
            )
        for key in ('word', 'meaning', 'hint1', 'hint2'):
            if key in request.data and request.data[key] is not None:
                setattr(word, key, request.data[key])
        word.is_approved = True
        word.approved_by = request.user
        word.assigned_to = None
        word.save(update_fields=['word', 'meaning', 'hint1', 'hint2', 'is_approved', 'approved_by', 'assigned_to'])
        CheckerAction.objects.update_or_create(
            word=word, user=request.user,
            defaults={'action': 'updated_and_approved'}
        )
        return Response({'status': 'success', 'message': 'Word updated and approved'})


class CheckWordsRejectView(APIView):
    """POST /api/checkwords/reject/ - reject word."""
    permission_classes = [IsAuthenticated, IsWordChecker]

    def post(self, request):
        word_id = request.data.get('word_id')
        if not word_id:
            return Response({'status': 'error', 'message': 'word_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            word = Word.objects.get(pk=word_id)
        except Word.DoesNotExist:
            return Response({'status': 'error', 'message': 'Word not found'}, status=status.HTTP_404_NOT_FOUND)
        if word.assigned_to_id is not None and word.assigned_to_id != request.user.id:
            return Response(
                {'status': 'error', 'message': 'You are not assigned to this word'},
                status=status.HTTP_403_FORBIDDEN,
            )
        word.is_rejected = True
        word.rejected_by = request.user
        word.assigned_to = None
        word.save(update_fields=['is_rejected', 'rejected_by', 'assigned_to'])
        CheckerAction.objects.update_or_create(
            word=word, user=request.user,
            defaults={'action': 'rejected'}
        )
        return Response({'status': 'success', 'message': 'Word rejected'})


class CheckWordsRemoveView(APIView):
    """POST /api/checkwords/remove/ - remove for me (skip)."""
    permission_classes = [IsAuthenticated, IsWordChecker]

    def post(self, request):
        word_id = request.data.get('word_id')
        if not word_id:
            return Response({'status': 'error', 'message': 'word_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            word = Word.objects.get(pk=word_id)
        except Word.DoesNotExist:
            return Response({'status': 'error', 'message': 'Word not found'}, status=status.HTTP_404_NOT_FOUND)
        if word.assigned_to_id is not None and word.assigned_to_id != request.user.id:
            return Response(
                {'status': 'error', 'message': 'You are not assigned to this word'},
                status=status.HTTP_403_FORBIDDEN,
            )
        word.assigned_to = None
        word.assigned_time = None
        word.save(update_fields=['assigned_to', 'assigned_time'])
        CheckerAction.objects.update_or_create(
            word=word, user=request.user,
            defaults={'action': 'removed_for_me'}
        )
        return Response({'status': 'success', 'message': 'Removed for you'})


class CheckWordsMyActionsView(APIView):
    """GET /api/checkwords/my-actions/?updated_after=<ts> - words I checked."""
    permission_classes = [IsAuthenticated, IsWordChecker]

    def get(self, request):
        qs = CheckerAction.objects.filter(user=request.user).select_related('word').order_by('-action_at')
        updated_after = request.query_params.get('updated_after')
        if updated_after:
            dt = parse_datetime(updated_after)
            if dt:
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
                qs = qs.filter(action_at__gt=dt)
        words = []
        for ca in qs:
            words.append({
                'id': ca.word_id,
                'word': ca.word.word,
                'meaning': ca.word.meaning or '',
                'status': ca.action,
                'assigned_time': ca.word.assigned_time.isoformat() if ca.word.assigned_time else None,
                'updated_time': ca.action_at.isoformat() if ca.action_at else None,
            })
        return Response({'words': words})
